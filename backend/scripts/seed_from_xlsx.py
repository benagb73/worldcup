"""
scripts/seed_from_xlsx.py

Build the database from a human-edited xlsx workbook. The spreadsheet is the
source of truth for pre-tournament data: teams, players, venues, the fixture
list and the knockout-bracket scaffolding.

Foreign keys in the workbook use human-friendly codes (team_code, venue_name,
club_name) so you never need to look up numeric IDs while typing.

Usage:
    python -m scripts.seed_from_xlsx --init      # create a fresh template
    python -m scripts.seed_from_xlsx             # import data/worldcup.xlsx → DB
    python -m scripts.seed_from_xlsx --file path/to/other.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.db.connection import get_db, init_db


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = REPO_ROOT / "data" / "worldcup.xlsx"


# ---------------------------------------------------------------------------
# Sheet schemas — header order matters; first non-empty row defines order
# ---------------------------------------------------------------------------

SHEETS: dict[str, list[str]] = {
    "teams":   ["name", "code", "group_name", "flag_url", "world_rank"],
    "venues":  ["name", "city", "country", "capacity", "number_games"],
    "clubs":   ["name", "country", "league"],   # country is 3-letter code
    "players": ["team_code", "club_name", "name", "shirt_number", "position", "date_of_birth"],
    "matches": ["match_number", "stage", "group_name", "scheduled_at",
                "home_team_code", "away_team_code", "venue_name"],
    "bracket": ["stage", "slot", "home_seed_desc", "away_seed_desc"],
}

# Recognised sentinel values for a missing-club entry (case-insensitive)
UNATTACHED_TOKENS = {"unattached", "free agent", "free-agent", "no club", "none"}
UNKNOWN_TOKENS    = {"unknown", "tbd", "?", "n/a", "na"}


def _normalise_dob(v) -> str | None:
    """Accept Excel datetime cells, ISO strings, or 'DD/MM/YYYY' strings → ISO 'YYYY-MM-DD'."""
    if v is None:
        return None
    # Excel date cells come through as datetime
    if hasattr(v, "isoformat"):
        return v.date().isoformat() if hasattr(v, "date") else v.isoformat()[:10]
    s = str(v).strip()
    if not s:
        return None
    # Already ISO?
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    # DD/MM/YYYY or DD-MM-YYYY
    for sep in ("/", "-", "."):
        parts = s.split(sep)
        if len(parts) == 3 and len(parts[2]) == 4:
            d, m, y = parts
            try:
                return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
            except ValueError:
                pass
    return s   # last-resort: keep whatever the user typed

# Dropdown options for cells where it helps a lot
ENUMS: dict[tuple[str, str], list[str]] = {
    ("players", "position"): ["GK", "DEF", "MID", "FWD"],
    ("matches", "stage"):    ["group", "r32", "r16", "qf", "sf", "third_place", "final"],
    ("bracket", "stage"):    ["r32", "r16", "qf", "sf", "third_place", "final"],
    ("teams",   "group_name"): list("ABCDEFGHIJKL"),  # WC 2026 has 12 groups
}

# Example rows so the user sees the format. Group A (USA pot example)
EXAMPLES: dict[str, list[dict]] = {
    "teams": [
        {"name": "United States", "code": "USA", "group_name": "A",
         "flag_url": "https://flagcdn.com/w320/us.png"},
        {"name": "Mexico",        "code": "MEX", "group_name": "A",
         "flag_url": "https://flagcdn.com/w320/mx.png"},
        {"name": "Canada",        "code": "CAN", "group_name": "B",
         "flag_url": "https://flagcdn.com/w320/ca.png"},
    ],
    "venues": [
        {"name": "MetLife Stadium",   "city": "East Rutherford", "country": "USA",    "capacity": 82500},
        {"name": "AT&T Stadium",      "city": "Arlington",       "country": "USA",    "capacity": 80000},
        {"name": "BMO Field",         "city": "Toronto",         "country": "Canada", "capacity": 45000},
        {"name": "Estadio Azteca",    "city": "Mexico City",     "country": "Mexico", "capacity": 87000},
    ],
    "clubs": [
        {"name": "Real Madrid", "country": "Spain",   "league": "La Liga"},
        {"name": "Manchester City", "country": "England", "league": "Premier League"},
        {"name": "Inter Miami", "country": "USA", "league": "MLS"},
    ],
    "players": [
        {"team_code": "USA", "club_name": "Inter Miami", "name": "Christian Pulisic",
         "shirt_number": 10, "position": "FWD", "date_of_birth": "1998-09-18"},
        {"team_code": "MEX", "club_name": "",            "name": "Edson Alvarez",
         "shirt_number": 4,  "position": "MID", "date_of_birth": "1997-10-24"},
    ],
    "matches": [
        # match_number is your own ordering — used as a stable identifier so
        # admin-page updates can find the row even if the auto id changes.
        {"match_number": 1, "stage": "group", "group_name": "A",
         "scheduled_at": "2026-06-11T20:00:00Z",
         "home_team_code": "MEX", "away_team_code": "USA",
         "venue_name": "Estadio Azteca"},
    ],
    "bracket": [
        {"stage": "r32", "slot": 1,
         "home_seed_desc": "Winner Group A", "away_seed_desc": "3rd best 3rd-placed"},
        {"stage": "final", "slot": 1,
         "home_seed_desc": "Winner SF1",     "away_seed_desc": "Winner SF2"},
    ],
}


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------

def build_template(out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2A47")
    note_font   = Font(italic=True, color="666666")

    # README sheet first
    readme = wb.create_sheet("README", 0)
    notes = [
        ("World Cup 2026 — Data Workbook",),
        ("",),
        ("Edit each sheet, then run:  python -m scripts.seed_from_xlsx",),
        ("",),
        ("Sheets:",),
        ("  teams     — one row per nation. code is a 3-letter id used in matches/players.",),
        ("  venues    — host stadiums. Referenced by name from matches.",),
        ("  clubs     — domestic clubs. Referenced by name from players.",),
        ("  players   — squad lists. team_code links to teams.code, club_name to clubs.name.",),
        ("  matches   — full fixture list. Use the team and venue lookups above.",),
        ("  bracket   — knockout slots seeded ahead of the draw (e.g. 'Winner Group A').",),
        ("",),
        ("Tips:",),
        ("  • scheduled_at is an ISO-8601 UTC timestamp:  2026-06-11T20:00:00Z",),
        ("  • position is one of: GK / DEF / MID / FWD",),
        ("  • stage is: group / r32 / r16 / qf / sf / third_place / final",),
        ("  • Re-running the seeder wipes & repopulates ALL tables. Don't run it",),
        ("    once the tournament is live — use the /admin UI for in-flight updates.",),
    ]
    for r, (txt,) in enumerate(notes, 1):
        cell = readme.cell(row=r, column=1, value=txt)
        if r == 1:
            cell.font = Font(bold=True, size=16, color="C9A84C")
        elif txt.startswith(("Sheets:", "Tips:")):
            cell.font = Font(bold=True)
        else:
            cell.font = note_font
    readme.column_dimensions["A"].width = 88

    # Data sheets
    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # Drop in example rows
        for r, example in enumerate(EXAMPLES.get(sheet_name, []), start=2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=r, column=col, value=example.get(header, ""))

        # Add dropdown validations
        for (s_name, col_name), choices in ENUMS.items():
            if s_name != sheet_name:
                continue
            if col_name not in headers:
                continue
            col_idx = headers.index(col_name) + 1
            letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(choices) + '"',
                allow_blank=True,
            )
            dv.add(f"{letter}2:{letter}1000")
            ws.add_data_validation(dv)

    wb.save(out_path)
    print(f"[OK] Template written to {out_path}")
    print(f"     Edit each sheet, then run:  python -m scripts.seed_from_xlsx")


# ---------------------------------------------------------------------------
# Import: xlsx → DB
# ---------------------------------------------------------------------------

def _read_sheet(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result: list[dict] = []
    for row in rows[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            if isinstance(v, str):
                v = v.strip()
                if v == "":
                    v = None
            item[h] = v
        result.append(item)
    return result


async def import_workbook(xlsx_path: Path) -> None:
    if not xlsx_path.exists():
        print(f"[ERROR] Workbook not found: {xlsx_path}")
        print(f"        Run with --init to generate a starter template.")
        sys.exit(1)

    print(f"[1/3] Reading {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)

    teams    = _read_sheet(wb, "teams")
    venues   = _read_sheet(wb, "venues")
    clubs    = _read_sheet(wb, "clubs")
    players  = _read_sheet(wb, "players")
    matches  = _read_sheet(wb, "matches")
    bracket  = _read_sheet(wb, "bracket")

    print(f"      teams={len(teams)}  venues={len(venues)}  clubs={len(clubs)}")
    print(f"      players={len(players)}  matches={len(matches)}  bracket={len(bracket)}")

    print(f"\n[2/3] Initialising database")
    await init_db()

    print(f"\n[3/3] Writing rows")
    async with get_db() as db:

        # Wipe everything in FK-safe order (children first, parents last).
        # `picks` references both players AND matches, so it must go before either.
        # competitors / pools / pool_members / comp_scoring are user data — preserved.
        print("      Clearing existing data...")
        for table in ["picks",
                      "player_match_stats", "match_lineups", "match_events",
                      "bracket", "group_standings", "matches",
                      "players", "teams", "venues", "clubs"]:
            await db.execute(f"DELETE FROM {table}")

        # --- teams -----------------------------------------------------------
        team_id_by_code: dict[str, int] = {}
        for t in teams:
            if not t.get("name") or not t.get("code"):
                continue
            wr = t.get("world_rank")
            wr = int(wr) if wr not in (None, "") else None
            rows = await db.execute(
                "INSERT INTO teams (name, code, group_name, flag_url, world_rank) "
                "VALUES (?,?,?,?,?) RETURNING id",
                [t["name"], t["code"], t.get("group_name"), t.get("flag_url"), wr]
            )
            if rows:
                team_id_by_code[t["code"]] = rows[0]["id"]
        print(f"      teams        -> {len(team_id_by_code)} inserted")

        # --- venues ----------------------------------------------------------
        venue_id_by_name: dict[str, int] = {}
        for v in venues:
            if not v.get("name"):
                continue
            cap = v.get("capacity")
            cap = int(cap) if cap not in (None, "") else None
            ng  = v.get("number_games")
            ng  = int(ng) if ng not in (None, "") else None
            rows = await db.execute(
                "INSERT INTO venues (name, city, country, capacity, number_games) "
                "VALUES (?,?,?,?,?) RETURNING id",
                [v["name"], v.get("city") or "", v.get("country") or "", cap, ng]
            )
            if rows:
                venue_id_by_name[v["name"]] = rows[0]["id"]
        print(f"      venues       -> {len(venue_id_by_name)} inserted")

        # --- clubs -----------------------------------------------------------
        club_id_by_name: dict[str, int] = {}
        for c in clubs:
            if not c.get("name"):
                continue
            rows = await db.execute(
                "INSERT INTO clubs (name, country, league) VALUES (?,?,?) RETURNING id",
                [c["name"], c.get("country") or "", c.get("league") or ""]
            )
            if rows:
                club_id_by_name[c["name"]] = rows[0]["id"]
        print(f"      clubs        -> {len(club_id_by_name)} inserted")

        # --- players ---------------------------------------------------------
        n_players = 0
        n_skipped_players = 0
        n_unattached = n_unknown = 0
        warned_clubs: set[str] = set()
        for p in players:
            if not p.get("name") or not p.get("team_code"):
                n_skipped_players += 1
                continue
            tid = team_id_by_code.get(p["team_code"])
            if not tid:
                print(f"      [WARN] player {p['name']!r}: unknown team_code {p['team_code']!r}")
                n_skipped_players += 1
                continue

            # Resolve club, handling sentinel values and unknown names
            raw_club = (p.get("club_name") or "").strip()
            lower_club = raw_club.lower()
            cid: int | None = None
            club_status: str | None = None
            if not raw_club:
                pass   # blank → no club info, no badge
            elif lower_club in UNATTACHED_TOKENS:
                club_status = "unattached"
                n_unattached += 1
            elif lower_club in UNKNOWN_TOKENS:
                club_status = "unknown"
                n_unknown += 1
            else:
                cid = club_id_by_name.get(raw_club)
                if cid is None and raw_club not in warned_clubs:
                    print(f"      [WARN] player {p['name']!r}: club {raw_club!r} not in clubs sheet (player inserted with no club)")
                    warned_clubs.add(raw_club)

            shirt = p.get("shirt_number")
            shirt = int(shirt) if shirt not in (None, "") else None
            dob   = _normalise_dob(p.get("date_of_birth"))

            await db.execute(
                "INSERT INTO players (team_id, club_id, name, shirt_number, position, date_of_birth, club_status) "
                "VALUES (?,?,?,?,?,?,?)",
                [tid, cid, p["name"], shirt, p.get("position"), dob, club_status]
            )
            n_players += 1
        print(f"      players      -> {n_players} inserted ({n_skipped_players} skipped) "
              f"[{n_unattached} unattached, {n_unknown} unknown club]")

        # --- matches ---------------------------------------------------------
        # Group-stage matches require both teams. Knockout matches are allowed
        # to be inserted with NULL teams — they're placeholders that the admin
        # UI fills in via "Set teams" once teams qualify.
        n_matches = n_ko_placeholders = 0
        n_skipped_matches = 0
        for m in matches:
            stage = (m.get("stage") or "group").strip()
            home_code = (m.get("home_team_code") or "").strip()
            away_code = (m.get("away_team_code") or "").strip()
            hid = team_id_by_code.get(home_code) if home_code else None
            aid = team_id_by_code.get(away_code) if away_code else None

            # Group stage must have both teams resolved
            if stage == "group":
                if not hid or not aid:
                    print(f"      [WARN] match #{m.get('match_number')} ({stage}): "
                          f"unknown / missing team ({home_code!r} vs {away_code!r})")
                    n_skipped_matches += 1
                    continue
            else:
                # Knockouts: tolerate missing teams (placeholder), but warn if a
                # code was typed but didn't resolve.
                if home_code and not hid:
                    print(f"      [WARN] match #{m.get('match_number')} ({stage}): "
                          f"home_team_code {home_code!r} not in teams sheet")
                if away_code and not aid:
                    print(f"      [WARN] match #{m.get('match_number')} ({stage}): "
                          f"away_team_code {away_code!r} not in teams sheet")
                if not hid and not aid:
                    n_ko_placeholders += 1

            vid = venue_id_by_name.get(m.get("venue_name")) if m.get("venue_name") else None
            sched = m.get("scheduled_at")
            if sched is not None and not isinstance(sched, str):
                sched = sched.isoformat()
            match_num = m.get("match_number")
            match_num = int(match_num) if match_num not in (None, "") else None
            await db.execute(
                "INSERT INTO matches (stage, group_name, match_number, "
                "home_team_id, away_team_id, venue_id, scheduled_at, status) "
                "VALUES (?,?,?,?,?,?,?, 'scheduled')",
                [stage, m.get("group_name"), match_num,
                 hid, aid, vid, sched]
            )
            n_matches += 1
        print(f"      matches      -> {n_matches} inserted ({n_skipped_matches} skipped) "
              f"[{n_ko_placeholders} knockout placeholders awaiting teams]")

        # --- bracket slots ---------------------------------------------------
        for b in bracket:
            if not b.get("stage") or b.get("slot") in (None, ""):
                continue
            await db.execute(
                "INSERT INTO bracket (stage, slot, home_seed_desc, away_seed_desc) "
                "VALUES (?,?,?,?)",
                [b["stage"], int(b["slot"]), b.get("home_seed_desc"), b.get("away_seed_desc")]
            )
        print(f"      bracket      -> {len(bracket)} slots inserted")

        # --- group_standings (zeroed rows for every grouped team) ------------
        n_st = 0
        for code, tid in team_id_by_code.items():
            t = next((x for x in teams if x["code"] == code), None)
            if not t or not t.get("group_name"):
                continue
            await db.execute(
                "INSERT INTO group_standings (group_name, team_id) VALUES (?,?)",
                [t["group_name"], tid]
            )
            n_st += 1
        print(f"      standings    -> {n_st} zeroed rows inserted")

    print("\n[OK] Database populated from workbook.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

async def main():
    p = argparse.ArgumentParser(description="Seed the World Cup DB from an xlsx workbook.")
    p.add_argument("--init", action="store_true",
                   help="Generate a fresh template at data/worldcup.xlsx")
    p.add_argument("--file", type=Path, default=DEFAULT_XLSX,
                   help="Path to the workbook (default: data/worldcup.xlsx)")
    args = p.parse_args()

    if args.init:
        if args.file.exists():
            print(f"[ERROR] {args.file} already exists. Delete or rename it first.")
            sys.exit(1)
        build_template(args.file)
        return

    await import_workbook(args.file)


if __name__ == "__main__":
    asyncio.run(main())
