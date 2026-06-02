"""
scripts/wiki_export_xlsx.py

Scrape the Wikipedia "2026 FIFA World Cup squads" page and OVERWRITE the
clubs + players sheets of data/worldcup.xlsx with the parsed data.

The teams, venues, matches, bracket sheets are left untouched.

Usage:
    python -m scripts.wiki_export_xlsx                  # write to data/worldcup.xlsx
    python -m scripts.wiki_export_xlsx --out=tmp.xlsx   # write to a different file
    python -m scripts.wiki_export_xlsx --dry-run        # don't save, just print summary

The script reads team codes from the existing teams sheet. League is left
blank (the user fills those in manually). Club country is a 3-letter code
derived from the Wikipedia flag icon's alt text.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

# Force UTF-8 output for Windows console
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.sync_from_wikipedia import (
    fetch_page, parse_squads, _norm, TEAM_NAME_OVERRIDES,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = REPO_ROOT / "data" / "worldcup.xlsx"


# ---------------------------------------------------------------------------
# Country name → ISO 3166-1 alpha-3 code.
# Curated for football: covers the WC 2026 teams plus most clubs' countries.
# Falls back to the raw name if not found (user can fix in xlsx).
# ---------------------------------------------------------------------------

COUNTRY_CODES: dict[str, str] = {
    # WC 2026 participants (use these to match the user's teams sheet codes)
    "argentina": "ARG", "australia": "AUS", "austria": "AUT",
    "belgium": "BEL", "brazil": "BRA", "cabo verde": "CPV", "cape verde": "CPV",
    "canada": "CAN", "chile": "CHI", "colombia": "COL", "costa rica": "CRC",
    "croatia": "CRO", "czech republic": "CZE", "czechia": "CZE",
    "denmark": "DEN", "dominican republic": "DOM", "ecuador": "ECU",
    "egypt": "EGY", "england": "ENG", "fiji": "FIJ", "france": "FRA",
    "germany": "GER", "ghana": "GHA", "greece": "GRE", "haiti": "HAI",
    "honduras": "HON", "iceland": "ISL", "iran": "IRN", "iraq": "IRQ",
    "italy": "ITA", "ivory coast": "CIV", "côte d'ivoire": "CIV", "cote d'ivoire": "CIV",
    "jamaica": "JAM", "japan": "JPN", "jordan": "JOR", "kazakhstan": "KAZ",
    "korea republic": "KOR", "south korea": "KOR", "korea, south": "KOR",
    "mexico": "MEX", "morocco": "MAR", "netherlands": "NED",
    "new zealand": "NZL", "nigeria": "NGA", "northern ireland": "NIR",
    "norway": "NOR", "oman": "OMA", "panama": "PAN", "paraguay": "PAR",
    "peru": "PER", "poland": "POL", "portugal": "POR", "qatar": "QAT",
    "republic of ireland": "IRL", "ireland": "IRL",
    "saudi arabia": "KSA", "scotland": "SCO", "senegal": "SEN",
    "serbia": "SRB", "slovakia": "SVK", "slovenia": "SVN",
    "south africa": "RSA", "spain": "ESP", "sweden": "SWE",
    "switzerland": "SUI", "syria": "SYR", "tunisia": "TUN",
    "turkey": "TUR", "türkiye": "TUR", "turkiye": "TUR",
    "ukraine": "UKR", "united arab emirates": "UAE", "uae": "UAE",
    "united states": "USA", "uruguay": "URU", "uzbekistan": "UZB",
    "venezuela": "VEN", "wales": "WAL",
    # Other countries with significant pro leagues that show up in club rows
    "albania": "ALB", "algeria": "ALG", "armenia": "ARM", "azerbaijan": "AZE",
    "bahrain": "BHR", "bangladesh": "BAN", "belarus": "BLR",
    "bosnia and herzegovina": "BIH", "bulgaria": "BUL",
    "cambodia": "CAM", "cameroon": "CMR", "china": "CHN", "china pr": "CHN",
    "cyprus": "CYP", "dr congo": "COD", "democratic republic of the congo": "COD",
    "estonia": "EST", "ethiopia": "ETH",
    "finland": "FIN", "georgia": "GEO", "guinea": "GUI",
    "hungary": "HUN", "india": "IND", "indonesia": "IDN",
    "israel": "ISR", "kenya": "KEN", "kosovo": "KOS", "kuwait": "KUW",
    "latvia": "LVA", "lebanon": "LBN", "libya": "LBY", "lithuania": "LTU",
    "luxembourg": "LUX", "malaysia": "MAS", "malta": "MLT",
    "moldova": "MDA", "montenegro": "MNE", "mozambique": "MOZ",
    "north macedonia": "MKD", "macedonia": "MKD",
    "north korea": "PRK", "korea dpr": "PRK",
    "palestine": "PLE", "philippines": "PHI",
    "romania": "ROU", "russia": "RUS", "rwanda": "RWA",
    "singapore": "SIN", "thailand": "THA", "trinidad and tobago": "TRI",
    "uganda": "UGA", "vietnam": "VIE", "zambia": "ZAM", "zimbabwe": "ZIM",
}


# Wikipedia's flag icon `alt` text is the football federation name (e.g.
# "Royal Dutch Football Association"), not a country name. Map them to codes.
FEDERATION_CODES: dict[str, str] = {
    "algerian football federation":                       "ALG",
    "argentine football association":                     "ARG",
    "association of football federations of azerbaijan":  "AZE",
    "austrian football association":                      "AUT",
    "brazilian football confederation":                   "BRA",
    "bulgarian football union":                           "BUL",
    "canadian soccer association":                        "CAN",
    "chinese football association":                       "CHN",
    "colombian football federation":                      "COL",
    "costa rican football federation":                    "CRC",
    "croatian football federation":                       "CRO",
    "cyprus football association":                        "CYP",
    "danish football association":                        "DEN",
    "ecuadorian football federation":                     "ECU",
    "egyptian football association":                      "EGY",
    "football association of bosnia and herzegovina":     "BIH",
    "football association of finland":                    "FIN",
    "football association of indonesia":                  "IDN",
    "football association of ireland":                    "IRL",
    "football association of malaysia":                   "MAS",
    "football association of serbia":                     "SRB",
    "football association of slovenia":                   "SVN",
    "football association of thailand":                   "THA",
    "football association of wales":                      "WAL",
    "football association of the czech republic":         "CZE",
    "football australia":                                 "AUS",
    "football federation islamic republic of iran":       "IRN",
    "football federation of armenia":                     "ARM",
    "football federation of chile":                       "CHI",
    "french football federation":                         "FRA",
    "german football association":                        "GER",
    "ghana football association":                         "GHA",
    "haitian football federation":                        "HAI",
    "hellenic football federation":                       "GRE",
    "hungarian football federation":                      "HUN",
    "iraq football association":                          "IRQ",
    "israel football association":                        "ISR",
    "italian football federation":                        "ITA",
    "japan football association":                         "JPN",
    "jordan football association":                        "JOR",
    "kazakhstan football federation":                     "KAZ",
    "korea football association":                         "KOR",
    "mexican football federation":                        "MEX",
    "national autonomous federation of football of honduras": "HON",
    "new zealand football":                               "NZL",
    "norwegian football federation":                      "NOR",
    "panamanian football federation":                     "PAN",
    "paraguayan football association":                    "PAR",
    "polish football association":                        "POL",
    "portuguese football federation":                     "POR",
    "qatar football association":                         "QAT",
    "romanian football federation":                       "ROU",
    "royal belgian football association":                 "BEL",
    "royal dutch football association":                   "NED",
    "royal moroccan football federation":                 "MAR",
    "royal spanish football federation":                  "ESP",
    "russian football union":                             "RUS",
    "saudi arabian football federation":                  "KSA",
    "scottish football association":                      "SCO",
    "slovak football association":                        "SVK",
    "south african football association":                 "RSA",
    "swedish football association":                       "SWE",
    "swiss football association":                         "SUI",
    "the football association":                           "ENG",   # The FA = England
    "tunisian football federation":                       "TUN",
    "turkish football federation":                        "TUR",
    "united arab emirates football association":          "UAE",
    "united states soccer federation":                    "USA",
    "uruguayan football association":                     "URU",
    "uzbekistan football association":                    "UZB",
    "venezuelan football federation":                     "VEN",
}


def country_to_code(name: str | None) -> str:
    """Map either a country name OR a Wikipedia federation name to a 3-letter code."""
    if not name:
        return ""
    n = _norm(name)
    # Try federation lookup first (most cells deliver this), then country
    return FEDERATION_CODES.get(n) or COUNTRY_CODES.get(n) or name.strip()


# ---------------------------------------------------------------------------
# Sheet helpers (mirror seed_from_xlsx.py headers)
# ---------------------------------------------------------------------------

CLUB_HEADERS   = ["name", "country", "league"]
PLAYER_HEADERS = ["team_code", "club_name", "name", "shirt_number",
                  "position", "date_of_birth"]


def _ensure_sheet(wb, name: str, headers: list[str]):
    """Return a fresh sheet with the given headers (replaces existing sheet)."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2A47")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--out", type=Path, default=DEFAULT_XLSX,
                   help="Output xlsx path (default: data/worldcup.xlsx)")
    p.add_argument("--dry-run", action="store_true",
                   help="Don't save, just print a summary")
    args = p.parse_args()

    if not args.out.exists():
        print(f"[ERROR] {args.out} doesn't exist — create the template first via "
              f"`python -m scripts.seed_from_xlsx --init`")
        sys.exit(1)

    # Build team_name → team_code from the existing teams sheet
    wb = load_workbook(args.out)
    if "teams" not in wb.sheetnames:
        print("[ERROR] teams sheet missing from the workbook")
        sys.exit(1)
    teams_sheet = wb["teams"]
    team_rows = list(teams_sheet.iter_rows(values_only=True))
    if len(team_rows) < 2:
        print("[ERROR] teams sheet is empty — fill it in before running this")
        sys.exit(1)
    team_headers = [str(h).strip() if h is not None else "" for h in team_rows[0]]
    try:
        i_name = team_headers.index("name")
        i_code = team_headers.index("code")
    except ValueError:
        print(f"[ERROR] teams sheet needs 'name' and 'code' columns; got: {team_headers}")
        sys.exit(1)

    teamcode_by_norm: dict[str, str] = {}
    for row in team_rows[1:]:
        if row[i_name] and row[i_code]:
            teamcode_by_norm[_norm(str(row[i_name]))] = str(row[i_code]).strip()
            teamcode_by_norm[_norm(str(row[i_code]))] = str(row[i_code]).strip()

    print(f"Loaded {len(teamcode_by_norm)//2} teams from {args.out.name}")
    print("Fetching Wikipedia squads page…")
    html = fetch_page()
    teams = parse_squads(html)
    print(f"Parsed {len(teams)} team sections from Wikipedia\n")

    # Collect unique clubs (deduplicate by club_name) + flatten all players
    clubs: dict[str, str] = {}      # club_name → country code
    players_rows: list[tuple] = []
    unresolved_teams: list[str] = []
    unmapped_country_names: set[str] = set()
    skipped_no_club = 0

    for wt in teams:
        wiki_team_name = wt["team_name"]
        lookup = TEAM_NAME_OVERRIDES.get(_norm(wiki_team_name), _norm(wiki_team_name))
        team_code = teamcode_by_norm.get(lookup) or teamcode_by_norm.get(_norm(wiki_team_name))
        if not team_code:
            unresolved_teams.append(wiki_team_name)
            print(f"  [SKIP] '{wiki_team_name}' — no matching team_code in xlsx")
            continue

        for wp in wt["players"]:
            club_name = wp.get("club_name") or ""
            raw_country = wp.get("club_country") or ""
            country   = country_to_code(raw_country)
            # If lookup didn't reduce the string to a 3-letter code, flag it
            if raw_country and country == raw_country.strip():
                unmapped_country_names.add(raw_country)
            if club_name:
                # First occurrence wins (Wikipedia is consistent about club country)
                if club_name not in clubs:
                    clubs[club_name] = country
            else:
                skipped_no_club += 1

            players_rows.append((
                team_code,
                club_name,
                wp.get("name"),
                wp.get("shirt_number"),
                wp.get("position"),
                wp.get("dob"),
            ))

    print(f"\n--- Summary ---")
    print(f"  Players parsed:        {len(players_rows)}")
    print(f"  Unique clubs:          {len(clubs)}")
    print(f"  Teams skipped:         {len(unresolved_teams)} {unresolved_teams or ''}")
    print(f"  Players w/o club row:  {skipped_no_club}")
    if unmapped_country_names:
        print(f"  Country names not in COUNTRY_CODES dict ({len(unmapped_country_names)}):")
        for n in sorted(unmapped_country_names)[:15]:
            print(f"    - {n}")
        if len(unmapped_country_names) > 15:
            print(f"    ...and {len(unmapped_country_names) - 15} more")
        print(f"  These clubs will have the country name as-is instead of a 3-letter code.")

    if args.dry_run:
        print("\n[DRY RUN] Not writing the xlsx. Re-run without --dry-run to save.")
        return

    # Build fresh sheets
    print(f"\nWriting clubs + players sheets to {args.out}")
    ws_c = _ensure_sheet(wb, "clubs",   CLUB_HEADERS)
    for r, (name, country) in enumerate(sorted(clubs.items()), start=2):
        ws_c.cell(row=r, column=1, value=name)
        ws_c.cell(row=r, column=2, value=country)
        # column 3 = league, left blank

    ws_p = _ensure_sheet(wb, "players", PLAYER_HEADERS)
    # Sort by team_code, then shirt number for sanity
    players_rows.sort(key=lambda x: (x[0] or "", x[3] if x[3] is not None else 99, x[2] or ""))
    for r, row in enumerate(players_rows, start=2):
        for c, val in enumerate(row, start=1):
            ws_p.cell(row=r, column=c, value=val)

    try:
        wb.save(args.out)
    except PermissionError:
        print(f"\n[ERROR] Couldn't write to {args.out} — close Excel and try again.")
        sys.exit(1)
    print(f"[OK] Wrote {len(clubs)} clubs and {len(players_rows)} players to {args.out}")


if __name__ == "__main__":
    main()
